# Импортируем необходимые библиотеки
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

# Генератор входных значений OCEAN (от 0 до 1)
def generate_ocean_values(num_samples):
    for _ in range(num_samples):
        yield {
            'openness': np.round(np.random.uniform(0, 1), 6),
            'conscientiousness': np.round(np.random.uniform(0, 1), 6),
            'extraversion': np.round(np.random.uniform(0, 1), 6),
            'agreeableness': np.round(np.random.uniform(0, 1), 6),
            'neuroticism': np.round(np.random.uniform(0, 1), 6)
        }

# Функция для расчета MBTI шкал на основе коэффициентов для каждой составляющей
def calculate_mbti_scales(ocean_values):
    openness = ocean_values['openness']
    conscientiousness = ocean_values['conscientiousness']
    extraversion = ocean_values['extraversion']
    agreeableness = ocean_values['agreeableness']
    neuroticism = ocean_values['neuroticism']
    
    # (a) MBTI_EI
    EI = (-0.01 * neuroticism) + (0.16 * extraversion) + (0.00 * openness) + (0.02 * agreeableness) + (-0.01 * conscientiousness)
    
    # (b) MBTI_SN
    SN = (-0.02 * neuroticism) + (-0.01 * extraversion) + (-0.03 * openness) + (0.05 * agreeableness) + (-0.05 * conscientiousness)
    
    # (c) MBTI_TF
    TF = (-0.12 * neuroticism) + (-0.04 * extraversion) + (-0.08 * openness) + (-0.39 * agreeableness) + (0.17 * conscientiousness)
    
    # (d) MBTI_JP
    JP = (0.04 * neuroticism) + (-0.02 * extraversion) + (-0.11 * openness) + (0.03 * agreeableness) + (0.13 * conscientiousness)
    
    return round(EI, 3), round(SN, 3), round(TF, 3), round(JP, 3)

# Функция для определения типа MBTI на основе полученных шкал
def mbti_type(EI, SN, TF, JP):
    ei_type = 'E' if EI >= 0.08 else 'I'
    sn_type = 'S' if SN >= -0.03 else 'N'
    tf_type = 'T' if TF >= -0.23 else 'F'
    jp_type = 'J' if JP >= 0.04 else 'P'
    return ei_type + sn_type + tf_type + jp_type

# Списки для хранения значений EI, SN, TF, JP и подготовка данных для записи в Excel
data = []
num_samples = 500000  # Количество значений для генерации

# Генерация и сбор данных в список для каждой строки
for ocean_values in generate_ocean_values(num_samples):
    EI, SN, TF, JP = calculate_mbti_scales(ocean_values)
    mbti = mbti_type(EI, SN, TF, JP)
    
    # Добавляем строку в данные
    data.append({
        "openness": ocean_values['openness'],
        "conscientiousness": ocean_values['conscientiousness'],
        "extraversion": ocean_values['extraversion'],
        "agreeableness": ocean_values['agreeableness'],
        "neuroticism": ocean_values['neuroticism'],
        "EI": EI,
        "SN": SN,
        "TF": TF,
        "JP": JP,
        "MBTI Type": mbti
    })

# Создаем DataFrame
df = pd.DataFrame(data)

# Вычисление среднего значения для каждого показателя
average_EI = round(df["EI"].mean(), 3)
average_SN = round(df["SN"].mean(), 3)
average_TF = round(df["TF"].mean(), 3)
average_JP = round(df["JP"].mean(), 3)

# Создаем новый Excel файл
wb = Workbook()
ws = wb.active

# ЗЗаписываем данные, оставляя строку
for row in dataframe_to_rows(df, index=False, header=True):
    ws.append(row)

# Оставляем строку пустой
ws.append([""] * len(df.columns))

# Добавляем строку со средними значениями и выделяем её зелёным цветом
average_row = ["Average ", "", "", "", "", average_EI, average_SN, average_TF, average_JP, ""]
ws.append(average_row)

# Задаем зелёную заливку
green_fill = PatternFill(start_color="B7DA92", end_color="B7DA92", fill_type="solid")
for cell in ws[ws.max_row]:  # ws.max_row указывает на последний ряд
    cell.fill = green_fill

# Устанавливаем ширину всех столбцов на 18 пикселей (около 3,5 см)
for col in ws.columns:
    max_length = 18
    ws.column_dimensions[col[0].column_letter].width = max_length

# Сохраняем Excel файл
wb.save("MBTI_Report.xlsx")
print("Отчёт успешно сохранен в файл MBTI_Report.xlsx")